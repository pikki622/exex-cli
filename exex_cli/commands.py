import cleo

from exex import parse
from openpyxl import load_workbook

from exex_cli import convert
from exex_cli import formats


class ExtractCommand(cleo.Command):
    """
    Extract data from Excel document.

    extract
        {filename : Source file path}
        {--s|sheet=0 : Name of sheet}
        {--r|range=all : Range}
        {--f|format=text : text, table, json, csv}
        {--d|delimiter=, : Delimiter in output}
    """

    def handle(self):
        # file
        arg_filename = self.argument("filename")
        book = load_workbook(arg_filename)

        # sheet
        arg_sheet = self.option("sheet")

        if arg_sheet == "0":
            arg_sheet = book.sheetnames[0]

        sheet = book[arg_sheet]

        # values
        arg_range = self.option("range")

        if arg_range == "all":
            values_raw = sheet.values
        else:
            values_raw = sheet[arg_range]

        values_parsed = parse.values(values_raw)

        # format
        arg_format = self.option("format")
        values_formatted = self.__format(values_parsed, arg_format)

        # render
        self.info("\n")
        self.__render(values_formatted, arg_format)
        self.info("\n")

        return

    def __format(self, values, format_):
        if format_ == formats.TEXT:
            return values
        elif format_ == formats.CSV:
            return convert.to_csv(values, delimiter=self.option("delimiter"))
        elif format_ == formats.TABLE:
            return values
        elif format_ == formats.JSON:
            return convert.to_json(values)
        else:
            return values

    def __render(self, values_formatted, format_):
        if format_ == formats.TABLE:
            self.render_table(headers=values_formatted[0], rows=values_formatted[1:])
        else:
            self.info(values_formatted)
